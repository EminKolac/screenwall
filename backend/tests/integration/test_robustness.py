"""Negative / robustness tests from the Codex Phase-2 review (H2/H3/M3/M5)."""
from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.config import get_settings
from app.extraction.base import UploadRejected
from app.extraction.xlsx import XlsxExtractor
from app.main import app
from app.models.document import DocumentStatus, Language
from app.services.ingest import ingest

client = TestClient(app)
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


def test_corrupt_ooxml_rejected_fail_closed():
    # Valid zip magic but not a real OOXML package.
    fake = b"PK\x03\x04" + b"\x00" * 64
    with pytest.raises(UploadRejected):
        ingest(fake, "evil.docx", get_settings())


def test_empty_pdf_routes_to_human_review():
    import fitz

    doc = fitz.open()
    doc.new_page()  # blank, no text
    data = doc.tobytes()
    doc.close()
    d, content = ingest(data, "scan.pdf", get_settings())
    assert d.status == DocumentStatus.NEEDS_HUMAN_REVIEW
    assert d.language == Language.unknown
    assert any("scanned" in w or "no extractable text" in w for w in content.warnings)


def test_sparse_xlsx_does_not_explode():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Ahmet Yılmaz"
    ws.cell(row=5000, column=20, value="far")  # far sparse cell
    buf = io.BytesIO()
    wb.save(buf)
    c = XlsxExtractor().extract(buf.getvalue(), "s.xlsx")
    text = c.plain_text  # must not allocate a dense 5000x20 grid
    assert "Ahmet Yılmaz" in text and "far" in text


def test_api_unsupported_type_415():
    r = client.post("/api/documents", files={"file": ("a.txt", b"hello world", "text/plain")})
    assert r.status_code == 415


def test_api_corrupt_docx_400():
    r = client.post(
        "/api/documents",
        files={"file": ("a.docx", b"PK\x03\x04" + b"\x00" * 64, _DOCX_MIME)},
    )
    assert r.status_code == 400
