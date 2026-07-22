"""Build an XLSX carrier embedding canaries across XLSX-specific channels: normal cell, cell
comment, hidden sheet, merged cell, and a line-broken cell. Marker in column A, value in column B
(same row) so the extracted row 'RefNNNN | value' lets the harness locate each value. Returns
(bytes, filename, placements)."""
from __future__ import annotations

import io

import openpyxl
from openpyxl.comments import Comment

from evaluation.bist30.canary import catalog_by_id
from evaluation.bist30.harness import Markers, Placement, make_placement


def build_xlsx(mk: Markers) -> tuple[bytes, str, list[Placement]]:
    cat = catalog_by_id()
    placements: list[Placement] = []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finansal Özet"
    row = 1

    def put(cid: str, channel: str, variant: str = "", sheet=None):
        nonlocal row
        p = make_placement(mk, cat[cid], "xlsx", channel, variant)
        tgt = sheet or ws
        tgt.cell(row=row, column=1, value=p.marker)
        tgt.cell(row=row, column=2, value=p.value)
        placements.append(p)
        r = row
        row += 1
        return p, r

    # Normal cells — a spread of critical types.
    for cid in ("prs_tr", "gsm_tr", "eml", "iban", "card", "tckn", "acct", "secret"):
        put(cid, "cell")
    # Line-broken value inside a single cell.
    put("iban", "cell", "line_break")

    # Cell comment — the canary lives in the comment text, with the marker inside it.
    pc = make_placement(mk, cat["tckn"], "xlsx", "comment")
    ws.cell(row=row, column=1, value=pc.marker)
    ws.cell(row=row, column=1).comment = Comment(pc.carrier_text(), "QA")
    placements.append(pc)
    row += 1

    # Merged cell — value in the top-left of a merged range (only the anchor holds the value).
    pm = make_placement(mk, cat["eml"], "xlsx", "merged")
    ws.cell(row=row, column=1, value=pm.marker)
    ws.cell(row=row, column=2, value=pm.value)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    placements.append(pm)
    row += 1

    # Hidden sheet — content that a user does not see but that is still in the workbook.
    hidden = wb.create_sheet("Gizli")
    hidden.sheet_state = "hidden"
    hrow = 1
    for cid in ("iban", "secret"):
        p = make_placement(mk, cat[cid], "xlsx", "hidden_sheet")
        hidden.cell(row=hrow, column=1, value=p.marker)
        hidden.cell(row=hrow, column=2, value=p.value)
        placements.append(p)
        hrow += 1

    # Filename embeds a person-name canary.
    fp = make_placement(mk, cat["prs_tr"], "xlsx", "filename")
    placements.append(fp)
    filename = f"Finansal Tablolar {fp.value} {fp.marker}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename, placements
