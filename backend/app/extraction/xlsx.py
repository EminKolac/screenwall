"""XLSX extraction via openpyxl. Each sheet becomes one table block with typed cells (row/col +
A1 address), preceded by a HEADING block carrying the sheet name so that name is anonymized +
audited too (rendering the raw `Block.sheet` would expose an un-audited name). Cell COMMENTS are
captured as well so their PII is audited + rendered (Codex P1).

Memory safety: VALUES are read with `read_only=True` streaming + a hard cell cap. COMMENTS are read
straight from the `xl/comments*.xml` parts of the zip — but each part is size-bounded BEFORE
decompression/parse (a single part could otherwise be hundreds of MB under the archive limit).
`data_only=True` yields the cached formula result (the audited value), not the formula source.
"""
from __future__ import annotations

import io
import zipfile
from xml.etree import ElementTree as ET

import openpyxl

from app.extraction.base import Block, BlockType, ExtractedContent, TableCell
from app.models.document import FileKind

_MAX_CELLS = 200_000
_MAX_COMMENTS = 20_000
_MAX_COMMENT_PART_BYTES = 5 * 1024 * 1024    # bound one comments part before decompress + parse
_MAX_COMMENT_TOTAL_BYTES = 20 * 1024 * 1024  # bound the SUM across all comment parts
_SS_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def _iter_comment_texts(data: bytes, cap: int):
    """Yield (ref, text) from xl/comments*.xml without loading the worksheet grid (bounded)."""
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception:  # noqa: BLE001
        return
    n = 0
    budget = 0
    with zf:
        for info in zf.infolist():
            name = info.filename
            low = name.lower()
            if not (name.startswith("xl/") and "comments" in low and low.endswith(".xml")):
                continue
            if info.file_size > _MAX_COMMENT_PART_BYTES:  # bound one part before read/parse
                continue
            budget += info.file_size
            if budget > _MAX_COMMENT_TOTAL_BYTES:  # bound the sum across parts
                break
            try:
                root = ET.fromstring(zf.read(name))
            except Exception:  # noqa: BLE001
                continue
            for c in root.iter(f"{_SS_NS}comment"):
                text = "".join(t.text or "" for t in c.iter(f"{_SS_NS}t")).strip()
                if text:
                    yield c.get("ref", ""), text
                    n += 1
                    if n >= cap:
                        return


class XlsxExtractor:
    kind = FileKind.xlsx

    def extract(self, data: bytes, filename: str) -> ExtractedContent:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        blocks: list[Block] = []
        warnings: list[str] = []
        count = 0
        truncated = False
        try:
            for ws in wb.worksheets:
                if truncated:
                    break
                cells: list[TableCell] = []
                for row in ws.iter_rows():
                    if truncated:
                        break
                    for cell in row:
                        if cell.value is None:
                            continue
                        if count >= _MAX_CELLS:
                            truncated = True
                            break
                        cells.append(TableCell(
                            row=cell.row - 1, col=cell.column - 1,
                            text=str(cell.value), address=cell.coordinate))
                        count += 1
                if cells:
                    # Sheet name as a heading block → anonymized + audited (not rendered raw).
                    blocks.append(Block(
                        block_id=f"sheet-h-{ws.title}", type=BlockType.heading,
                        text=ws.title, location=f"sheet {ws.title}"))
                    blocks.append(Block(
                        block_id=f"sheet-{ws.title}", type=BlockType.table,
                        cells=cells, sheet=ws.title, location=f"sheet {ws.title}"))
            meta = {"sheets": ",".join(wb.sheetnames)}
        finally:
            wb.close()

        # Comments (read from the zip XML — bounded, memory-safe).
        for i, (ref, text) in enumerate(_iter_comment_texts(data, _MAX_COMMENTS)):
            blocks.append(Block(
                block_id=f"cmt-{i}", type=BlockType.paragraph,
                text=f"{ref}: {text}" if ref else text, location="comment"))

        if truncated:
            warnings.append(f"workbook truncated at {_MAX_CELLS} cells")
        return ExtractedContent(kind=FileKind.xlsx, blocks=blocks, metadata=meta, warnings=warnings)
